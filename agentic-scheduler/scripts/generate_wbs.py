# -*- coding: utf-8 -*-
"""日本式 WBS＋日別ガント Excel 生成器（agentic-scheduler 附属）

用法:  python scripts/generate_wbs.py spec.json
       パスはスキルディレクトリ基準。このリポジトリ（スキルの monorepo）で動かす
       場合は先に `cd agentic-scheduler`。インストール済みスキルとして使う場合は
       そのスキルディレクトリが基準になる。
规格:  见 references/wbs-rendering.md（JSON schema 与渲染规则）

内置的既知坑位（勿改动除非重新验证）:
- 条件格式 dxf 的填充必须用 bgColor **且不要指定 patternType**
  （出力は <patternFill><bgColor/></patternFill>。これが CF dxf の規範形。
   solid を足すと fgColor 側が有効になり、未指定の fgColor のせいで
   「白字・白背景の隠身列」になる——実測で踏んだ罠。指摘されても戻さないこと）
- 当日红线用 CF 挂 TODAY()，打开文件自动追随；纯 CF 无单元格公式风险
- 父行/分组行条形用 lightTrellis 格子纹，与叶子实心条区分（计数一致性）
- 動的シート只用 2007 安全函数（IF/AND/COUNTIF/SUMPRODUCT/INDEX/MATCH/IFERROR/
  TODAY/ROW）——禁 FILTER/XLOOKUP/SORT（LibreOffice 不能算 + openpyxl 无 spill 元数据）
- 日別シート名は跨年時のみ %y%m%d（%m%d だと 12/30 と翌年 12/30 が衝突し、
  openpyxl が黙って "12301" のような紛らわしい名前に改名する）
- 連番＝終了日昇順 rank（SUMPRODUCT，同日按行序破并列保证 MATCH 唯一）
"""
import io
import json
import sys
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter

JP = "Yu Gothic"
BAR = "5B9BD5"    # 実装・調査（濃青）
SOFT = "BDD7EE"   # 検証・調整（薄青）
MS = "ED7D31"     # マイルストーン ◆（橙）
LV1 = "D9E2F3"    # 親行帯（情報列のみ）
GRAY = "EDEDED"   # 土日
HDR = "1F4E79"
STYLE_MAP = {"plan": BAR, "verify": SOFT}

INFO_HEADERS = ["WBS", "タスク", "種別", "Issue", "依存",
                "工数目安\n(稼働日)", "開始\n(目安)", "終了\n(目安)", "状態", "本日"]
N_INFO = len(INFO_HEADERS)          # 本日列 = 第 10 列（J）
COL0 = N_INFO + 1                   # 日列起点 = K
YOBI = "月火水木金土日"


def d(s):
    y, m, dd = (int(x) for x in s.split("-"))
    return date(y, m, dd)


def has_child(wbs, all_wbs):
    return any(o.startswith(wbs + ".") for o in all_wbs if o != wbs)


def build(spec):
    start = d(spec["start"])
    ndays = int(spec["days"])
    tasks = spec["tasks"]
    # 早期検証。ここを通すと R1 < R0 になって参照範囲が逆転し、openpyxl の奥から
    # 「expected MultiCellRange」という spec の誤りを示唆しない例外が飛ぶ。
    if ndays < 1:
        raise ValueError('spec の "days" は 1 以上にしてください（現在: {}）。'.format(ndays))
    if not tasks:
        raise ValueError('spec の "tasks" が空です。最低 1 行必要です。')
    days = [start + timedelta(i) for i in range(ndays)]
    all_wbs = [str(t["wbs"]) for t in tasks]
    n_rows = len(tasks)
    R0 = 4
    R1 = R0 + n_rows - 1
    cS, cE, cN = COL0 + ndays, COL0 + ndays + 1, COL0 + ndays + 2  # 隠し: 開始/終了/連番

    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    f_hdr = PatternFill("solid", fgColor=HDR)
    f_lv1 = PatternFill("solid", fgColor=LV1)
    f_gray = PatternFill("solid", fgColor=GRAY)
    solid = {c: PatternFill("solid", fgColor=c) for c in (BAR, SOFT, MS)}
    mesh = {c: PatternFill(patternType="lightTrellis", fgColor=c) for c in (BAR, SOFT)}

    wb = Workbook()
    ws = wb.active
    ws.title = "WBS"

    # ---- 表头（信息列 3 行合并 + 月/日/曜日）----
    for i, h in enumerate(INFO_HEADERS, 1):
        ws.merge_cells(start_row=1, start_column=i, end_row=3, end_column=i)
        c = ws.cell(1, i, h)
        c.font = Font(name=JP, bold=True, color="FFFFFF", size=10)
        c.fill = f_hdr
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    col = COL0
    runs = []
    for dd_ in days:
        if runs and runs[-1][0] == dd_.month:
            runs[-1][2] = col
        else:
            runs.append([dd_.month, col, col])
        col += 1
    for m, c1, c2 in runs:
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        c = ws.cell(1, c1, str(m) + "月")
        c.font = Font(name=JP, bold=True, color="FFFFFF", size=9)
        c.fill = f_hdr
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, dd_ in enumerate(days):
        cd = ws.cell(2, COL0 + i, dd_)
        cd.number_format = "d"
        cy = ws.cell(3, COL0 + i, YOBI[dd_.weekday()])
        for c in (cd, cy):
            c.font = Font(name=JP, size=8)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if dd_.weekday() >= 5:
                c.fill = f_gray
    ws.cell(3, cS, "開始日")
    ws.cell(3, cE, "終了日")
    ws.cell(3, cN, "連番")

    # ---- 任务行 ----
    JCOL = get_column_letter(N_INFO)                     # 本日列（J）
    SL, EL = get_column_letter(cS), get_column_letter(cE)
    daily_rows = []  # (row, 開始date, 終了date) — daily_sheets 用
    for idx, t in enumerate(tasks):
        r = R0 + idx
        w = str(t["wbs"])
        level = w.count(".")
        bold = level == 0
        grp = has_child(w, all_wbs)
        vals = [w, ("　" * level) + t["name"], t.get("kind", ""), t.get("issue", ""),
                t.get("dep", ""), t.get("effort", ""), t.get("start_note", ""),
                t.get("end_note", ""), t.get("status", "未着手")]
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v)
            c.font = Font(name=JP, size=10, bold=bold)
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=(i == 2))
            if bold:
                c.fill = f_lv1
        cj = ws.cell(r, N_INFO)
        cj.border = border
        cj.alignment = Alignment(horizontal="center", vertical="center")
        cj.font = Font(name=JP, size=10, bold=True, color="C00000")
        if bold:
            cj.fill = f_lv1
        if spec.get("task_dir") and not grp:
            ca = ws.cell(r, 1)
            ca.hyperlink = spec["task_dir"] + "/" + w + ".md"
            ca.font = Font(name=JP, size=10, bold=bold, color="0563C1", underline="single")
        # 土日底色
        for i, dd_ in enumerate(days):
            if dd_.weekday() >= 5:
                ws.cell(r, COL0 + i).fill = f_gray
        # 条形（bars: [{from,to,style}]）与 ◆（milestones: [date]）
        span = []
        for seg in t.get("bars", []):
            hexc = STYLE_MAP.get(seg.get("style", "plan"), BAR)
            a, b = d(seg["from"]), d(seg["to"])
            cur = a
            while cur <= b:
                i = (cur - start).days
                if 0 <= i < ndays:
                    # 格子柄かどうかは grp（子を持つか）だけで決める。bold（＝階層が
                    # トップレベル）を混ぜると、子を持たないトップレベル葉タスクが
                    # 格子柄で描かれるのに「本日」では葉として数えられ、凡例
                    #「格子柄＝親・グループ行、数えない」と矛盾する。
                    ws.cell(r, COL0 + i).fill = (mesh if grp else solid)[hexc]
                    span.append(i)
                cur += timedelta(1)
        for msd in t.get("milestones", []):
            i = (d(msd) - start).days
            if 0 <= i < ndays:
                c = ws.cell(r, COL0 + i, "◆")
                c.fill = solid[MS]
                c.font = Font(name=JP, size=9, bold=True, color="FFFFFF")
                c.alignment = Alignment(horizontal="center", vertical="center")
                span.append(i)
        # 隠し列は「開始〜終了の連続 1 区間」しか表現できない。バーが分断していたり
        # ◆ がバー期間の外にあると、min/max の間の空白日まで「本日」対象に数えられ、
        # 凡例（バー期間が本日を含む葉タスク）と食い違う workbook が黙って出来上がる。
        # 直せないので spec 側で行を分けてもらう。
        if (not grp) and span:
            uniq = sorted(set(span))
            if uniq[-1] - uniq[0] + 1 != len(uniq):
                holes = [start + timedelta(i)
                         for i in range(uniq[0], uniq[-1] + 1) if i not in set(uniq)]
                raise ValueError(
                    "WBS {0}「{1}」: バー／◆ の期間が連続していません（空白日: {2}）。"
                    "隠し列は連続 1 区間しか表せず、空白日まで「本日」対象に数えて"
                    "しまいます。行を分けるか、◆ を独立した行にしてください。".format(
                        w, t["name"].strip(),
                        "、".join(x.strftime("%m/%d") for x in holes[:5])
                        + ("…" if len(holes) > 5 else "")))
            ws.cell(r, cS, start + timedelta(min(span))).number_format = "m/d"
            ws.cell(r, cE, start + timedelta(max(span))).number_format = "m/d"
            daily_rows.append((r, start + timedelta(min(span)), start + timedelta(max(span))))
        if not grp:
            cj.value = ('=IF(AND(${S}{r}<>"",${S}{r}<=TODAY(),${E}{r}>=TODAY()),"●","")'
                        .format(S=SL, E=EL, r=r))
        ws.cell(r, cN,
                ('=IF(${J}{r}="●",'
                 'SUMPRODUCT((${J}${R0}:${J}${R1}="●")*(${E}${R0}:${E}${R1}<${E}{r}))'
                 '+SUMPRODUCT((${J}${R0}:${J}${R1}="●")*(${E}${R0}:${E}${R1}=${E}{r})'
                 '*(ROW(${E}${R0}:${E}${R1})<{r}))+1,"")')
                .format(J=JCOL, E=EL, r=r, R0=R0, R1=R1))

    note = ("※ 塗りつぶしバー＝葉タスク（「本日」列と本日のタスクの対象）。"
            "格子柄バー＝親行・グループ行の期間帯（数えない）。◆＝マイルストーン。"
            "赤い縦線＝本日（自動追随）。灰色列＝土日。日付は週粒度の目安。")
    ws.cell(R1 + 2, 2, note).font = Font(name=JP, size=9, color="777777")

    # ---- 当日红线（CF・dxf は bgColor！）----
    first, lastc = get_column_letter(COL0), get_column_letter(COL0 + ndays - 1)
    ws.conditional_formatting.add(
        "{0}2:{1}3".format(first, lastc),
        Rule(type="expression", formula=[first + "$2=TODAY()"],
             dxf=DifferentialStyle(
                 font=Font(name=JP, size=8, bold=True, color="FFFFFF"),
                 fill=PatternFill(bgColor="E74C3C"))))
    ws.conditional_formatting.add(
        "{0}{1}:{2}{3}".format(first, R0, lastc, R1),
        Rule(type="expression", formula=[first + "$2=TODAY()"],
             dxf=DifferentialStyle(border=Border(
                 left=Side(style="thick", color="FF0000"),
                 right=Side(style="thin", color="FF9999")))))

    for i, wd in enumerate([7, 56, 11, 11, 11, 9, 8, 10, 7, 5], 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    for i in range(ndays):
        ws.column_dimensions[get_column_letter(COL0 + i)].width = 3.0
    for c in (cS, cE, cN):
        ws.column_dimensions[get_column_letter(c)].hidden = True
    ws.freeze_panes = get_column_letter(COL0) + "4"
    ws.row_dimensions[1].height = 18
    ws.sheet_view.zoomScale = 80

    # ---- 本日のタスク（動的・終了日昇順）----
    t2 = wb.create_sheet("本日のタスク")
    t2["A1"] = "本日のタスク（ファイルを開いた日に自動更新・終了目安の早い順）"
    t2["A1"].font = Font(name=JP, size=12, bold=True)
    t2["A3"] = "本日："
    t2["B3"] = "=TODAY()"
    t2["B3"].number_format = "yyyy/m/d"
    t2["D3"] = "対象件数："
    t2["E3"] = '=COUNTIF(WBS!${J}${R0}:${J}${R1},"●")'.format(J=JCOL, R0=R0, R1=R1)
    for ref in ("A3", "D3"):
        t2[ref].font = Font(name=JP, size=10, bold=True)
    for ref in ("B3", "E3"):
        t2[ref].font = Font(name=JP, size=10)
    for i, h in enumerate(["#", "WBS", "タスク", "種別", "開始(目安)", "終了(目安)", "状態"], 1):
        c = t2.cell(5, i, h)
        c.font = Font(name=JP, size=10, bold=True, color="FFFFFF")
        c.fill = f_hdr
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    srcmap = {2: "A", 3: "B", 4: "C", 5: SL, 6: EL, 7: "I"}
    slots = int(spec.get("today_slots", 20))
    NL = get_column_letter(cN)
    for n in range(1, slots + 1):
        r = 5 + n
        t2.cell(r, 1, n).font = Font(name=JP, size=10, color="AAAAAA")
        t2.cell(r, 1).alignment = Alignment(horizontal="center")
        for col_i, src in srcmap.items():
            fml = ('=IFERROR(INDEX(WBS!${s}${R0}:${s}${R1},'
                   'MATCH($A{r},WBS!${N}${R0}:${N}${R1},0)),"")'
                   .format(s=src, r=r, N=NL, R0=R0, R1=R1))
            c = t2.cell(r, col_i, fml)
            c.font = Font(name=JP, size=10)
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=(col_i == 3))
            if col_i in (5, 6):
                c.number_format = "m/d"
    t2.cell(5 + slots + 2, 1,
            "※ 対象＝バー期間が本日を含む葉タスク（◆は当日のみ）。親行・グループ行は含まない"
            "——WBS シートの「本日」列 ● と同じ集合。終了(目安)の早い順（同日は WBS 順）。空欄＝該当なし。"
            ).font = Font(name=JP, size=9, color="777777")
    for i, wd in enumerate([4, 8, 62, 12, 10, 10, 8], 1):
        t2.column_dimensions[get_column_letter(i)].width = wd
    t2.freeze_panes = "A6"

    if spec.get("daily_sheets"):
        _add_daily_sheets(wb, days, daily_rows, SL, EL, f_hdr, border)

    # ---- 前提・凡例 ----
    ws3 = wb.create_sheet("前提・凡例")
    rows = [[spec.get("title", "WBS") + " — 前提・凡例", ""], ["", ""]]
    if spec.get("notes"):
        rows.append(["前提", ""])
        rows += [[k, v] for k, v in spec["notes"]]
        rows.append(["", ""])
    rows += [
        ["バーの凡例", ""],
        ["濃い青（塗り）", "設計・実装・調査の期間（葉タスク）"],
        ["薄い青（塗り）", "検証・調整の期間（暦日を要する・葉タスク）"],
        ["格子柄バー", "親行・グループ行の期間帯。作業そのものではないため「本日」列・本日のタスクには数えない"],
        ["橙 ◆", "マイルストーン（本番投入・判定など。当日のみ対象）"],
        ["赤い縦線", "本日（ファイルを開いた日に自動追随）"],
        ["灰色列", "土日"],
        ["「本日」列 ●", "バー期間が本日を含む葉タスクの印。本日のタスク シートと同じ集合"],
    ]
    for row in rows:
        ws3.append(row)
    for row in ws3.iter_rows():
        for c in row:
            c.font = Font(name=JP, size=10)
            c.alignment = Alignment(vertical="center", wrap_text=True)
    ws3["A1"].font = Font(name=JP, size=12, bold=True)
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 90

    wb.save(spec["out"])
    return spec["out"], n_rows, ndays


def _add_daily_sheets(wb, days, daily_rows, SL, EL, f_hdr, border):
    """日別シート（1 日 1 枚・内容は WBS への参照）。tab 数は日数分増える点に注意。

    この定義は build() より後・__main__ ブロックより前に置くこと。__main__ の後ろに
    置くとスクリプト実行時にまだ定義されておらず、daily_sheets 指定で NameError になる。
    """
    WKND = PatternFill("solid", fgColor="F5E9E9")
    # 跨年する計画では %m%d が衝突する（12/30 と翌年 12/30）。衝突すると openpyxl は
    # 黙って "12301" のような日付に見えて日付でない名前に改名するので、跨年時だけ年を足す。
    fmt = "%m%d" if days[0].year == days[-1].year else "%y%m%d"
    if len(days) > 60:
        print("warning: 日別シートを {} 枚作ります。タブが多すぎて実用に耐えない場合は "
              "daily_sheets を外してください。".format(len(days)), file=sys.stderr)
    for i, day in enumerate(days):
        t = wb.create_sheet(day.strftime(fmt), 1 + i)
        t.cell(1, 1, "{}/{}（{}）のタスク".format(day.month, day.day, YOBI[day.weekday()])).font = Font(name=JP, size=12, bold=True)
        t.cell(2, 1, "内容は WBS シートからの参照。行構成は計画時点のもの——計画を変えたら再生成する。").font = Font(name=JP, size=8, color="777777")
        for ci, h in enumerate(["#", "WBS", "タスク", "種別", "開始(目安)", "終了(目安)", "状態"], 1):
            c = t.cell(4, ci, h)
            c.font = Font(name=JP, size=10, bold=True, color="FFFFFF")
            c.fill = f_hdr
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        todays = sorted([x for x in daily_rows if x[1] <= day <= x[2]], key=lambda x: (x[2], x[0]))
        t.cell(3, 1, "対象 {} 件".format(len(todays))).font = Font(name=JP, size=9, color="555555")
        rr = 5
        for r, s, e in todays:
            t.cell(rr, 1, rr - 4).font = Font(name=JP, size=10, color="AAAAAA")
            refs = {2: "=WBS!A{}", 3: "=WBS!B{}", 4: "=WBS!C{}",
                    5: "=WBS!" + SL + "{}", 6: "=WBS!" + EL + "{}", 7: "=WBS!I{}"}
            for ci, f in refs.items():
                c = t.cell(rr, ci, f.format(r))
                c.font = Font(name=JP, size=10)
                c.border = border
                c.alignment = Alignment(vertical="center", wrap_text=(ci == 3))
                if ci in (5, 6):
                    c.number_format = "m/d"
            if e == day:
                t.cell(rr, 6).font = Font(name=JP, size=10, bold=True, color="C00000")
            if day.weekday() >= 5:
                for ci in range(1, 8):
                    t.cell(rr, ci).fill = WKND
            rr += 1
        for ci, wd_ in enumerate([4, 9, 62, 13, 10, 10, 8], 1):
            t.column_dimensions[get_column_letter(ci)].width = wd_
        t.freeze_panes = "A5"


if __name__ == "__main__":
    # 端末が cp932 でも日本語を出せるようにする。import 時ではなくここで行うのは、
    # モジュールとして import された時（wbs-rendering.md の「python 側シミュレーション」
    # による検証がこれに当たる）に呼び出し側の stdout を勝手に差し替えないため。
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    if len(sys.argv) != 2:
        print("usage: python scripts/generate_wbs.py spec.json"
              "  (スキルディレクトリから。この monorepo では先に cd agentic-scheduler)")
        sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as f:
        spec = json.load(f)
    out, n, nd = build(spec)
    print("generated:", out, "| tasks:", n, "| days:", nd)
