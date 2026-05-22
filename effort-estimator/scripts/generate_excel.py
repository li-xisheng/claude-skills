"""
宸ユ暟瑕嬬 Excel 鐢熸垚銈广偗銉儣銉?鈥?姹庣敤銉诲啀鍒╃敤鍙兘

Usage:
    from generate_excel import EstimateWorkbook
    wb = EstimateWorkbook()
    wb.set_meta(title="妗堜欢鍚?, date="2026-05-19", config={...})
    # add sections -> add tasks -> finalize
    wb.add_section("A", "銉椼儹銈搞偋銈儓鍩虹洡")
    wb.add_task("A1", "kintone鎶€琛撴瑷?, o=2, m=4, p=8, note="鏈€鍎厛")
    wb.add_subtotal("A", "A. 銉椼儹銈搞偋銈儓鍩虹洡 灏忚▓")
    ...
    wb.finalize()
    wb.add_phase_sheet(phase1=[...], phase2=[...])
    wb.add_assumptions_sheet([("浠畾", "褰遍熆"), ...])
    wb.save("output.xlsx")
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# 鈹€鈹€ Styling constants (customize per project if needed) 鈹€鈹€

class Styles:
    """Centralized style definitions. Override per-project by passing a subclass."""
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
    SECTION_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    SECTION_FONT = Font(bold=True, size=11, color='1F4E79')
    SUBTOTAL_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    SUMMARY_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    INPUT_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    INPUT_FONT = Font(bold=True, color='C00000')
    FINAL_FONT = Font(bold=True, size=14, color='C00000')
    TITLE_FONT = Font(bold=True, size=14)
    SUBTITLE_FONT = Font(size=9, color='666666')
    BOLD_FONT = Font(bold=True)
    BOLD_FONT_L = Font(bold=True, size=11)
    CHECK_FONT = Font(size=9, color='006100')
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    CENTER = Alignment(horizontal='center')
    WRAP = Alignment(wrap_text=True)


class EstimateWorkbook:
    """
    Generic estimation workbook builder.

    Sheet structure:
        1. WBS+PERT 鈥?task breakdown with formulas, adjustable coefficients, summary
        2. Phase鍒嗗壊 (optional) 鈥?phase split view referencing sheet 1
        3. 浠畾涓€瑕?(optional) 鈥?assumptions table
    """

    def __init__(self, styles=None):
        self.wb = openpyxl.Workbook()
        self.styles = styles or Styles()
        self.meta = {}
        self.config = {}
        self.sections = []          # [(letter, label)]
        self.tasks = []             # [(type, *args)]
        self.section_task_rows = {} # letter -> [row_numbers]
        self.subtotal_rows = {}     # letter -> row_number
        self._row = 5               # current write row in sheet 1
        self._sheet1 = self.wb.active
        self._sheet1.title = 'WBS+PERT'
        self._finalized = False

    # 鈹€鈹€ Metadata & config 鈹€鈹€

    def set_meta(self, title, subtitle='', config=None):
        """
        Set workbook-level metadata.

        config dict keys (all optional):
            skill_coef (float): default 1.0
            mgmt_rate (float): default 0.20
            risk_rate (float): default 0.15
            skill_note (str): hint for skill coeff cell
            mgmt_note (str): hint for mgmt rate cell
            risk_note (str): hint for risk rate cell
            col_widths (dict): {col_letter: width}
        """
        self.meta = {'title': title, 'subtitle': subtitle}
        self.config = {
            'skill_coef': 1.0,
            'mgmt_rate': 0.20,
            'risk_rate': 0.15,
            'skill_note': 'Senior:0.8~1.0 Mid:1.2~1.5 Junior:2.0~3.0',
            'mgmt_note': '瀹氫緥銉汇儸銉撱儱銉笺兓銉夈偔銉ャ儭銉炽儓 10~20%',
            'risk_note': 'PERT娓涘崐銉兗銉? 30%鈫?5%',
            **(config or {})
        }

    # 鈹€鈹€ Sheet 1: WBS+PERT 鈹€鈹€

    def _setup_sheet1_columns(self):
        widths = self.config.get('col_widths', {})
        defaults = {'A': 6, 'B': 48, 'C': 10, 'D': 10, 'E': 10, 'F': 12, 'G': 16, 'H': 14}
        for col, w in {**defaults, **widths}.items():
            self._sheet1.column_dimensions[col].width = w

    def _write_sheet1_header(self):
        ws = self._sheet1
        self._setup_sheet1_columns()
        ws.merge_cells('A1:H1')
        ws['A1'] = self.meta.get('title', '宸ユ暟瑕嬬')
        ws['A1'].font = self.styles.TITLE_FONT
        if self.meta.get('subtitle'):
            ws.merge_cells('A2:G2')
            ws['A2'] = self.meta['subtitle']
            ws['A2'].font = self.styles.SUBTITLE_FONT
        headers = ['ID', '銈裤偣銈?, 'O (妤借Τ)', 'M (鏈€鍙兘)', 'P (鎮茶Τ)', 'PERT鏈熷緟鍊?, '鍌欒€?, '鍒嗘暎 蟽虏']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER

    def add_section(self, label):
        """Add a section header row. label is displayed text (e.g. '銆怉. 銉椼儹銈搞偋銈儓鍩虹洡銆?)."""
        self.tasks.append(('section', label))

    def add_task(self, tid, name, o, m, p, note=''):
        """Add a task row with O/M/P values. PERT formula = (O+4M+P)/6 auto-applied.
        tid must start with the section letter (e.g. 'A1', 'B2')."""
        if not (tid and tid[0].isalpha()):
            raise ValueError(
                f"Task ID '{tid}' must start with a section letter (e.g. 'A1', 'B2'). "
                "This is required for subtotal row grouping."
            )
        self.tasks.append(('task', tid, name, o, m, p, note))

    def add_subtotal(self, letter, label):
        """Add a subtotal row that SUMs all tasks under the given section letter."""
        self.tasks.append(('subtotal', letter, label))

    def _write_tasks(self):
        ws = self._sheet1
        for item in self.tasks:
            if item[0] == 'section':
                ws.merge_cells(f'A{self._row}:G{self._row}')
                ws[f'A{self._row}'] = item[1]
                ws[f'A{self._row}'].font = self.styles.SECTION_FONT
                ws[f'A{self._row}'].fill = self.styles.SECTION_FILL
                self._row += 1
                continue

            if item[0] == 'subtotal':
                _, letter, label = item
                if letter in self.section_task_rows:
                    fr = self.section_task_rows[letter][0]
                    lr = self.section_task_rows[letter][-1]
                    formula = f'=SUM(F{fr}:F{lr})'
                else:
                    formula = '0'
                ws.cell(row=self._row, column=1, value=letter).font = self.styles.BOLD_FONT
                ws.cell(row=self._row, column=2, value=label).font = self.styles.BOLD_FONT
                cell = ws.cell(row=self._row, column=6, value=formula)
                cell.font = self.styles.BOLD_FONT
                cell.number_format = '0.00'
                # Subtotal variance sum
                var_cell = ws.cell(row=self._row, column=8)
                var_cell.value = f'=SUM(H{fr}:H{lr})'
                var_cell.number_format = '0.0000'
                var_cell.font = self.styles.BOLD_FONT
                for c in range(1, 9):
                    ws.cell(row=self._row, column=c).fill = self.styles.SUBTOTAL_FILL
                    ws.cell(row=self._row, column=c).border = self.styles.THIN_BORDER
                self.subtotal_rows[letter] = self._row
                self._row += 1
                continue

            # task row
            _, tid, name, o, m, p, note = item
            ws.cell(row=self._row, column=1, value=tid)
            ws.cell(row=self._row, column=2, value=name).alignment = self.styles.WRAP
            ws.cell(row=self._row, column=3, value=o).alignment = self.styles.CENTER
            ws.cell(row=self._row, column=4, value=m).alignment = self.styles.CENTER
            ws.cell(row=self._row, column=5, value=p).alignment = self.styles.CENTER
            cell = ws.cell(row=self._row, column=6)
            cell.value = f'=(C{self._row}+4*D{self._row}+E{self._row})/6'
            cell.number_format = '0.00'
            cell.alignment = self.styles.CENTER
            ws.cell(row=self._row, column=7, value=note)
            # Column H: per-task variance 蟽虏 = ((P-O)/6)虏
            var_cell = ws.cell(row=self._row, column=8)
            var_cell.value = f'=((E{self._row}-C{self._row})/6)^2'
            var_cell.number_format = '0.0000'
            var_cell.alignment = self.styles.CENTER
            # Track section
            sec = tid[0] if tid else ''
            if sec not in self.section_task_rows:
                self.section_task_rows[sec] = []
            self.section_task_rows[sec].append(self._row)
            for c in range(1, 9):
                ws.cell(row=self._row, column=c).border = self.styles.THIN_BORDER
            self._row += 1

    def _write_summary(self, summary_sections, summary_labels):
        """summary_sections: list of section letters. summary_labels: {letter: label}."""
        ws = self._sheet1
        self._row += 1
        ws.merge_cells(f'A{self._row}:G{self._row}')
        ws[f'A{self._row}'] = '鈻?宸ョ▼鍒?PERT 绱斿伐鏁?闆嗚▓'
        ws[f'A{self._row}'].font = Font(bold=True, size=12)
        self._row += 1

        for sec in summary_sections:
            if sec in self.subtotal_rows:
                ws.merge_cells(f'B{self._row}:D{self._row}')
                ws.cell(row=self._row, column=2, value=summary_labels.get(sec, sec))
                cell = ws.cell(row=self._row, column=6, value=f'=F{self.subtotal_rows[sec]}')
                cell.number_format = '0.00'
                for c in range(1, 9):
                    ws.cell(row=self._row, column=c).border = self.styles.THIN_BORDER
                self._row += 1

        # Total PERT
        self._pert_total_row = self._row
        ws.merge_cells(f'B{self._row}:D{self._row}')
        ws.cell(row=self._row, column=2, value='PERT 绱斿伐鏁?鍚堣▓').font = self.styles.BOLD_FONT_L
        refs = '+'.join([f'F{self.subtotal_rows[s]}' for s in summary_sections if s in self.subtotal_rows])
        if not refs:
            refs = '0'
        cell = ws.cell(row=self._row, column=6, value=f'=({refs})')
        cell.number_format = '0.00'
        cell.font = self.styles.BOLD_FONT_L
        for c in range(1, 9):
            ws.cell(row=self._row, column=c).fill = self.styles.SUMMARY_FILL
            ws.cell(row=self._row, column=c).border = self.styles.THIN_BORDER
        self._row += 2

    def _write_adjustments(self):
        """Write the adjustable coefficient cells and the 4-step calculation."""
        ws = self._sheet1
        cfg = self.config

        ws.merge_cells(f'A{self._row}:H{self._row}')
        ws[f'A{self._row}'] = '鈻?瑾挎暣瑷堢畻'
        ws[f'A{self._row}'].font = Font(bold=True, size=12)
        self._row += 1

        # -- Adjustable input cells --
        inputs = [
            ('skill_coef',  '鎶€鑳戒總鏁?(澶夋洿鍙?鈫?',    cfg['skill_coef'], cfg['skill_note']),
            ('mgmt_rate',   '绠＄悊宸ユ暟鐜?(澶夋洿鍙?鈫?',   cfg['mgmt_rate'],  cfg['mgmt_note']),
            ('risk_rate',   '銉偣銈儛銉冦儠銈＄巼 (澶夋洿鍙?鈫?', cfg['risk_rate'],  cfg['risk_note']),
        ]
        cell_refs = {}
        for key, label, default, note in inputs:
            ws.cell(row=self._row, column=2, value=label)
            cell = ws.cell(row=self._row, column=3, value=default)
            cell.alignment = self.styles.CENTER
            cell.fill = self.styles.INPUT_FILL
            cell.font = self.styles.INPUT_FONT
            if isinstance(default, float) and default < 1:
                cell.number_format = '0%'
            ws.cell(row=self._row, column=7, value=note)
            for c in range(1, 9):
                ws.cell(row=self._row, column=c).border = self.styles.THIN_BORDER
            cell_refs[key] = f'C{self._row}'
            self._row += 1
        # Store for cross-sheet reference (e.g. Phase sheet)
        self._adjustment_cell_refs = cell_refs

        # -- Calculated steps: each step references the previous step's output row --
        # step_start + 0 = Step 1, step_start + 1 = Step 2, etc.
        step_start = self._row
        steps = [
            ('Step 1: PERT 绱斿伐鏁?,                         f'=F{self._pert_total_row}',                               '鍏ㄣ偪銈广偗銇笁鐐硅绌嶉泦瑷?),
            ('Step 2: 鎶€鑳戒總鏁拌鏁村緦',                       f'=F{step_start}*{cell_refs["skill_coef"]}',               ''),
            ('Step 3: 绠＄悊宸ユ暟鍔犵畻寰?,                       f'=F{step_start+1}*(1+{cell_refs["mgmt_rate"]})',           ''),
            ('Step 4: 銉偣銈儛銉冦儠銈″姞绠楀緦 鈽呮渶绲傚伐鏁扳槄',     f'=F{step_start+2}*(1+{cell_refs["risk_rate"]})',           ''),
        ]
        step_rows = {}
        for label, formula, note in steps:
            ws.cell(row=self._row, column=2, value=label)
            cell = ws.cell(row=self._row, column=6, value=formula)
            cell.number_format = '0.00'
            if '鈽? in label:
                cell.font = self.styles.FINAL_FONT
                for c in range(1, 9):
                    ws.cell(row=self._row, column=c).fill = self.styles.SUMMARY_FILL
            ws.cell(row=self._row, column=7, value=note)
            for c in range(1, 9):
                ws.cell(row=self._row, column=c).border = self.styles.THIN_BORDER
            if 'Step 4' in label:
                step_rows['step4'] = self._row
            self._row += 1

        # Anti-double-buffer check 鈥?actual Excel formula
        self._row += 1
        risk_ref = cell_refs.get("risk_rate", "0%")
        ws.merge_cells(f'A{self._row}:F{self._row}')
        ws.cell(row=self._row, column=1,
                value=f'鉁?Anti-double-buffer: PERT浣跨敤鈫掋儛銉冦儠銈℃笡鍗?| 鏈€绲傘儛銉冦儠銈＄巼=').font = self.styles.CHECK_FONT
        ws.cell(row=self._row, column=7).value = f'={risk_ref}'
        ws.cell(row=self._row, column=7).number_format = '0%'
        check = ws.cell(row=self._row, column=8)
        check.value = f'=IF({risk_ref}>0.5,"鈿?50%瓒?閲嶈鍙兘鎬?,"鉁?50%鏈簚 OK")'
        check.font = self.styles.CHECK_FONT

        # Add CI section with 3 range types
        self._write_ci_section(cell_refs, step_rows)

    def _write_ci_section(self, cell_refs, step_rows):
        """Add PERT CI section with 3 range types: absolute extremes + statistical CI + expected value."""
        ws = self._sheet1
        self._row += 2
        ws.merge_cells(f'A{self._row}:H{self._row}')
        ws.cell(row=self._row, column=1, value='鈻?PERT 淇￠牸鍖洪枔 (Confidence Interval)').font = Font(bold=True, size=12)
        self._row += 1

        ws.merge_cells(f'A{self._row}:H{self._row}')
        ws.cell(row=self._row, column=1, value='鍏∣鍊?鍏≒鍊?= 绲跺鐨勪笂涓嬮檺 | PERT绲辫▓CI = 鍚則ask鍋忓樊銇浉浜掓墦娑堛仐銈掕€冩叜').font = Font(size=9, color='666666')
        self._row += 1

        skill_ref = cell_refs.get("skill_coef", "C1")
        mgmt_ref = cell_refs.get("mgmt_rate", "C1")
        risk_ref = cell_refs.get("risk_rate", "C1")
        step4_row = step_rows['step4']
        adj_mul = f'{skill_ref}*(1+{mgmt_ref})*(1+{risk_ref})'
        final_ref = f'F{step4_row}'

        # Sigma formulas
        sigma_refs = [f'H{r}' for r in self.subtotal_rows.values()]
        sigma_formula = '=SQRT(' + '+'.join(sigma_refs) + ')' if sigma_refs else '=SQRT(0)'

        all_o_refs = [f'C{r}' for rows in self.section_task_rows.values() for r in rows]
        all_p_refs = [f'E{r}' for rows in self.section_task_rows.values() for r in rows]
        all_o_formula = '(' + '+'.join(all_o_refs) + ')'
        all_p_formula = '(' + '+'.join(all_p_refs) + ')'

        ws.cell(row=self._row, column=2, value='PERT 妯欐簴鍋忓樊 蟽_total')
        ws.cell(row=self._row, column=6, value=sigma_formula).number_format = '0.00'
        ws.cell(row=self._row, column=6).font = self.styles.BOLD_FONT
        ws.cell(row=self._row, column=7, value='鈭?危蟽虏_i)')
        for c in range(1, 9):
            ws.cell(row=self._row, column=c).border = self.styles.THIN_BORDER
        sigma_row = self._row
        self._row += 1

        ws.cell(row=self._row, column=2, value='瑾挎暣寰屾婧栧亸宸?(蟽_adjusted)')
        ws.cell(row=self._row, column=6, value=f'=F{sigma_row}*{adj_mul}').number_format = '0.00'
        ws.cell(row=self._row, column=7, value='蟽_total 脳 瑾挎暣淇傛暟锛堟渶绲傚伐鏁般偣銈便兗銉伀鎻涚畻锛?)
        for c in range(1, 9):
            ws.cell(row=self._row, column=c).border = self.styles.THIN_BORDER
        sigma_adj_row = self._row
        self._row += 1

        # 鈶?Absolute extremes
        self._row += 1
        section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        section_font = Font(bold=True, size=10, color='1F4E79')
        ws.merge_cells(f'A{self._row}:H{self._row}')
        ws.cell(row=self._row, column=1, value='鈶?鍏ㄦソ瑕炽€滃叏鎮茶Τ (绲跺鐨勪笂涓嬮檺 鈥?鍏╰ask銇屽悓鏅傘伀鏈€鑹?鏈€鎮仺銇倠妤点倎銇︾█銇偙銉笺偣)')
        ws.cell(row=self._row, column=1).font = section_font
        ws.cell(row=self._row, column=1).fill = section_fill
        self._row += 1

        for label, formula, note in [
            ('鍏∣鍊ゅ悎瑷?(妤借Τ妤靛€?', f'={all_o_formula}*{adj_mul}', '鍏╰ask銇屾ソ瑕冲€ら€氥倞閫层倱銇犲牬鍚堛伄涓嬮檺'),
            ('鍏≒鍊ゅ悎瑷?(鎮茶Τ妤靛€?', f'={all_p_formula}*{adj_mul}', '鍏╰ask銇屾偛瑕冲€ら€氥倞闆ｈ埅銇椼仧鍫村悎銇笂闄?),
        ]:
            ws.cell(row=self._row, column=2, value=label)
            ws.cell(row=self._row, column=6, value=formula).number_format = '0.0'
            ws.cell(row=self._row, column=7, value=note)
            for c in range(1, 9):
                ws.cell(row=self._row, column=c).border = self.styles.THIN_BORDER
            self._row += 1

        # 鈶?Statistical CI
        self._row += 1
        ws.merge_cells(f'A{self._row}:H{self._row}')
        ws.cell(row=self._row, column=1, value='鈶?PERT 绲辫▓鐨勪俊闋煎尯闁?(鍚則ask銇亸宸亴鐙珛銉荤浉浜掓墦娑堛仐銇欍倠銇撱仺銈掕€冩叜)')
        ws.cell(row=self._row, column=1).font = section_font
        ws.cell(row=self._row, column=1).fill = section_fill
        self._row += 1

        for label, formula, note in [
            ('68% CI 涓嬮檺 (卤1蟽)', f'={final_ref}-F{sigma_adj_row}', '绱?/3銇⒑鐜囥仹銇撱伄绡勫洸鍐?),
            ('68% CI 涓婇檺 (卤1蟽)', f'={final_ref}+F{sigma_adj_row}', ''),
            ('95% CI 涓嬮檺 (卤2蟽) 鈽?, f'={final_ref}-2*F{sigma_adj_row}', '绱?5%銇⒑鐜囷紙鎺ㄥエ鎻愮ず绡勫洸锛?),
            ('95% CI 涓婇檺 (卤2蟽) 鈽?, f'={final_ref}+2*F{sigma_adj_row}', 'PERT鏈熷緟鍊?卤 2蟽_adjusted'),
        ]:
            ws.cell(row=self._row, column=2, value=label)
            cell = ws.cell(row=self._row, column=6, value=formula)
            cell.number_format = '0.0'
            if '鈽? in label:
                cell.font = self.styles.FINAL_FONT
            ws.cell(row=self._row, column=7, value=note)
            for c in range(1, 9):
                ws.cell(row=self._row, column=c).border = self.styles.THIN_BORDER
                if '鈽? in label:
                    ws.cell(row=self._row, column=c).fill = self.styles.SUMMARY_FILL
            self._row += 1

        # 鈶?Expected value
        self._row += 1
        ws.merge_cells(f'A{self._row}:H{self._row}')
        ws.cell(row=self._row, column=1, value='鈶?PERT 鏈熷緟鍊?(鏈€鍙兘鍊?鈥?鏈绌嶃伄鎻愮ず鍊?')
        ws.cell(row=self._row, column=1).font = section_font
        ws.cell(row=self._row, column=1).fill = section_fill
        self._row += 1

        ws.cell(row=self._row, column=2, value='PERT 鏈熷緟鍊?(鏈绌嶃伄鎻愮ず鍊?')
        cell = ws.cell(row=self._row, column=6, value=f'={final_ref}')
        cell.number_format = '0.0'
        cell.font = self.styles.FINAL_FONT
        for c in range(1, 9):
            ws.cell(row=self._row, column=c).fill = self.styles.SUMMARY_FILL
            ws.cell(row=self._row, column=c).border = self.styles.THIN_BORDER

    def finalize(self, summary_sections=None, summary_labels=None):
        """
        Finalize sheet 1: write header, tasks, summary, adjustments.
        Must be called after all add_section/add_task/add_subtotal calls.
        summary_sections: ordered list of section letters for the summary table.
        summary_labels: {letter: 'A. 宸ョ▼鍚?} dict for display.
        """
        self._write_sheet1_header()
        self._write_tasks()
        if summary_sections is None:
            summary_sections = list(self.subtotal_rows.keys())
        if summary_labels is None:
            summary_labels = {s: s for s in summary_sections}
        self._write_summary(summary_sections, summary_labels)
        self._write_adjustments()
        self._sheet1.freeze_panes = 'A5'
        self._finalized = True

    # 鈹€鈹€ Sheet 2: Phase 鍒嗗壊 鈹€鈹€

    def add_phase_sheet(self, phases, sheet_name='Phase鍒嗗壊',
                        mgmt_rate=None, risk_rate=None):
        """
        Add a phase split sheet.

        phases: list of dicts, each dict:
            {
                'name': 'Phase 1',
                'sections': [('A. 銉椼儹銈搞偋銈儓鍩虹洡', 'A', 1.0), ...]
                    # (display_label, section_letter, ratio) 鈥?ratio: 鍓插悎 (0.5 for 50%)
                'note': '8鏈堛儶銉兗銈圭洰妯?,
                'summary_note': '绱?00浜烘棩',  # optional
            }
        mgmt_rate, risk_rate: override sheet1's config if given
        """
        if not self._finalized:
            raise RuntimeError("Call finalize() before add_phase_sheet()")

        ws = self.wb.create_sheet(sheet_name)
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 22

        ws.merge_cells('A1:E1')
        ws['A1'] = 'Phase 鍒嗗壊鎻愭'
        ws['A1'].font = self.styles.TITLE_FONT

        for col, h in enumerate(['Phase', '宸ョ▼', 'PERT绱斿伐鏁?, '瑾挎暣寰屽伐鏁?, '鍌欒€?], 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER

        # Use cell references from WBS+PERT if available, else fall back to config values
        refs = getattr(self, '_adjustment_cell_refs', {})
        skill_ref = refs.get('skill_coef', None)
        mgmt_ref = refs.get('mgmt_rate', None)
        risk_ref = refs.get('risk_rate', None)
        if skill_ref:
            skill_ref = f"'WBS+PERT'!{skill_ref}"
        else:
            skill_ref = self.config.get('skill_coef', 1.0)
        if mgmt_ref:
            mgmt_ref = f"'WBS+PERT'!{mgmt_ref}"
        else:
            mgmt_ref = mgmt_rate if mgmt_rate is not None else self.config['mgmt_rate']
        if risk_ref:
            risk_ref = f"'WBS+PERT'!{risk_ref}"
        else:
            risk_ref = risk_rate if risk_rate is not None else self.config['risk_rate']

        r = 4
        phase_totals = []

        for phase in phases:
            start_r = r
            for label, sec, ratio in phase['sections']:
                ws.cell(row=r, column=1, value=phase['name'])
                ws.cell(row=r, column=2, value=label)
                if ratio != 1.0:
                    ws.cell(row=r, column=3, value=f"='WBS+PERT'!F{self.subtotal_rows[sec]}*{ratio}")
                else:
                    ws.cell(row=r, column=3, value=f"='WBS+PERT'!F{self.subtotal_rows[sec]}")
                ws.cell(row=r, column=3).number_format = '0.00'
                ws.cell(row=r, column=4, value=f'=C{r}*{skill_ref}*(1+{mgmt_ref})*(1+{risk_ref})')
                ws.cell(row=r, column=4).number_format = '0.00'
                ws.cell(row=r, column=5, value=phase.get('note', ''))
                for c in range(1, 6):
                    ws.cell(row=r, column=c).border = self.styles.THIN_BORDER
                r += 1

            # Phase subtotal
            ws.merge_cells(f'A{r}:B{r}')
            ws.cell(row=r, column=1, value=f'{phase["name"]} 鍚堣▓').font = self.styles.BOLD_FONT
            ws.cell(row=r, column=3, value=f'=SUM(C{start_r}:C{r-1})').number_format = '0.00'
            ws.cell(row=r, column=3).font = self.styles.BOLD_FONT
            ws.cell(row=r, column=4, value=f'=SUM(D{start_r}:D{r-1})').number_format = '0.00'
            ws.cell(row=r, column=4).font = self.styles.BOLD_FONT
            sn = phase.get('summary_note', '')
            ws.cell(row=r, column=5, value=sn)
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = self.styles.SUMMARY_FILL
                ws.cell(row=r, column=c).border = self.styles.THIN_BORDER
            phase_totals.append(r)
            r += 1

        # Grand total
        r += 1
        ws.cell(row=r, column=2, value='鍏ㄤ綋鍚堣▓').font = Font(bold=True, size=12)
        refs = '+'.join([f'D{pt}' for pt in phase_totals])
        ws.cell(row=r, column=4, value=f'={refs}').number_format = '0.00'
        ws.cell(row=r, column=4).font = Font(bold=True, size=12)

        ws.freeze_panes = 'A4'

    # 鈹€鈹€ Sheet 3: 浠畾涓€瑕?鈹€鈹€

    def add_assumptions_sheet(self, assumptions, sheet_name='浠畾涓€瑕?):
        """
        Add an assumptions/reference sheet.

        assumptions: list of (assumption_text, impact_if_changed) tuples.
        """
        ws = self.wb.create_sheet(sheet_name)
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 35

        ws.merge_cells('A1:C1')
        ws['A1'] = '涓昏浠畾銇ㄥ墠鎻愭潯浠?
        ws['A1'].font = self.styles.TITLE_FONT

        for col, h in enumerate(['#', '浠畾', '澶夊嫊銇椼仧鍫村悎銇奖闊?], 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER

        for i, (assumption, impact) in enumerate(assumptions, 1):
            ws.cell(row=i+3, column=1, value=f'#{i}')
            ws.cell(row=i+3, column=2, value=assumption).alignment = self.styles.WRAP
            ws.cell(row=i+3, column=2).font = self.styles.BOLD_FONT
            ws.cell(row=i+3, column=3, value=impact).alignment = self.styles.WRAP
            for c in range(1, 4):
                ws.cell(row=i+3, column=c).border = self.styles.THIN_BORDER

    # 鈹€鈹€ Save 鈹€鈹€

    def save(self, path):
        if not self._finalized:
            raise RuntimeError("Call finalize() before save()")
        self.wb.save(path)
        print(f'Saved: {path}')


# 鈹€鈹€ Convenience: generate from a simple dict structure 鈹€鈹€

def from_dict(project: dict, output_path: str):
    """
    Quick generation from a declarative dict.

    project = {
        'title': '妗堜欢鍚?宸ユ暟瑕嬬',
        'subtitle': '浣滄垚鏃? 2026-05-19 | ...',
        'config': {  # optional overrides
            'skill_coef': 1.0, 'mgmt_rate': 0.20, 'risk_rate': 0.15,
            'skill_note': '...', 'mgmt_note': '...', 'risk_note': '...',
        },
        'sections': [
            ('A', '銆怉. 宸ョ▼鍚嶃€?, [
                ('A1', '銈裤偣銈悕', o, m, p, '鍌欒€?),
                ...
            ]),
        ],
        'summary_sections': ['A','B',...],
        'summary_labels': {'A': 'A. xx', ...},
        'phases': [  # optional
            {'name': 'Phase 1', 'sections': [...], 'note': '...'},
        ],
        'assumptions': [  # optional
            ('浠畾1', '褰遍熆1'),
        ],
    }
    """
    wb = EstimateWorkbook()
    wb.set_meta(
        title=project['title'],
        subtitle=project.get('subtitle', ''),
        config=project.get('config')
    )
    for letter, label, tasks in project['sections']:
        wb.add_section(label)
        for t in tasks:
            tid, name, o, m, p = t[0], t[1], t[2], t[3], t[4]
            note = t[5] if len(t) > 5 else ''
            wb.add_task(tid, name, o, m, p, note)
        wb.add_subtotal(letter, f'{label.strip("銆愩€?)} 灏忚▓')

    wb.finalize(
        summary_sections=project.get('summary_sections'),
        summary_labels=project.get('summary_labels')
    )
    if 'phases' in project:
        wb.add_phase_sheet(project['phases'])
    if 'assumptions' in project:
        wb.add_assumptions_sheet(project['assumptions'])

    wb.save(output_path)
    return wb


if __name__ == '__main__':
    # Example usage (abstract 鈥?replace with real project data)
    project = {
        'title': '銈点兂銉椼儷妗堜欢 宸ユ暟瑕嬬',
        'subtitle': '浣滄垚鏃? 2026-XX-XX | 闁嬬櫤鍩烘簴: 涓€鑸枊鐧鸿€?,
        'config': {},
        'sections': [
            ('A', '銆怉. 銈点兂銉椼儷宸ョ▼銆?, [
                ('A1', '銈点兂銉椼儷銈裤偣銈?', 1, 2, 3, ''),
                ('A2', '銈点兂銉椼儷銈裤偣銈?', 2, 3, 5, '鍌欒€冧緥'),
            ]),
            ('B', '銆怋. 鍒ュ伐绋嬨€?, [
                ('B1', '鍒ャ偪銈广偗1', 1, 1.5, 3, ''),
            ]),
        ],
        'summary_sections': ['A', 'B'],
        'summary_labels': {'A': 'A. 銈点兂銉椼儷宸ョ▼', 'B': 'B. 鍒ュ伐绋?},
        'assumptions': [
            ('浠畾銇緥1', '澶夊嫊銇椼仧鍫村悎 卤X浜烘棩'),
            ('浠畾銇緥2', '澶夊嫊銇椼仧鍫村悎 +Y浜烘棩'),
        ],
    }
    from_dict(project, '/tmp/sample_estimate.xlsx')
